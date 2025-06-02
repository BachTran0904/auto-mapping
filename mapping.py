import sys
import asyncio
import io
from openpyxl import load_workbook
import json
import argparse
import dify_api
import requests


#Load json file
def load_mappings(attribute_json_path):
    try:
        with open(attribute_json_path, 'r', encoding='utf-8') as f:
            mappings = json.load(f)
        return mappings
    except Exception as e:
        print(f"Error loading attribute JSON: {str(e)}")
        raise

#global variables
processed_columns = {}
target_columns = {}
mappings_attribute = load_mappings("attribute.json")
mappings_header = load_mappings("attribute_header.json")
mappings_full = load_mappings("attribute_sheet.json")

filepath = "Raw.xlsx"
filepath2 = "Raw2.xlsx"
formpath = "Form.xlsx"

def process_dify_api_output(input, source):
    global fields_collumn
    header = []

    header_row = next(source.iter_rows(min_row=1, max_row=1, values_only=True))
    for cell in header_row:
        if cell is not None: header.append(cell)

    lines = input.split('\n')

    for line in lines:
        # Bỏ qua các dòng không phải dữ liệu
        if line.strip() == "I cannot find a suitable pair." or line.strip() == "I cannot find a suitable attribute.":
            continue

        # Dòng dạng: Category: Attribute: Value
        parts = line.split(': ')
        if len(parts) >= 4:
            # Loại bỏ khoảng trắng thừa ở mỗi phần
            category = parts[0].strip()
            attribute = parts[1].strip()
            origin = parts[2].strip()
            title = parts[3].strip()

            #if ((category, attribute) not in processed_columns) & (origin in header):
                # Nếu cặp (category, attribute) đã tồn tại, có thể log hoặc bỏ qua
            processed_columns[(category, attribute)] = [title, origin]

        else:
            # Nếu không đúng định dạng, có thể log hoặc bỏ qua
            print(f"Warning: line không đúng định dạng: {line}")

async def call_dify_api(source_sheet):
    header_row = next(source_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    filtered_headers = [cell for cell in header_row if cell is not None]

    # Asynchronously get the response using asyncio.to_thread for sync functions
    response = dify_api.get_answer_from_dify(filtered_headers, source_sheet.title)
    # Process the response after it's received
    process_dify_api_output(response, source_sheet)

async def process_sheets(workbook):
    sheets = workbook.sheetnames  # Get sheet names

    tasks = []  # To hold all the tasks
    for sheet_name in sheets:
        sheet = workbook[sheet_name]  # Get the actual sheet
        # Asynchronously call the API function for each sheet
        tasks.append(call_dify_api(sheet))

    # Wait for all tasks to complete
    await asyncio.gather(*tasks)

def find_data_sheets(source_wb):
    """Find all sheets with data (not empty) in the workbook"""
    data_sheets = []
    for sheet in source_wb:
        if sheet.max_row > 0:  # At least have header row
            data_sheets.append(sheet)
    return data_sheets
        
def get_source_columns(source_sheet):
    global processed_columns
    field_columns= {}
    header_row = next(source_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    for idx, header in enumerate(header_row, start=1):
        for (category, attribute), ((source_sheet_name, source_collumn)) in processed_columns.items():
            if header == source_collumn:
                field_columns[source_collumn] = idx
                break
    return field_columns

def get_target_columns(target_sheet):
    global processed_columns
    target_columns = {}
    target_header_row = next(target_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    for idx, header in enumerate(target_header_row, start=1):
        for category, attribute in processed_columns:
            if header == attribute:
                target_columns[attribute] = idx
                break
    return target_columns


def copy_data_to_target(source_sheet, target_sheet):
    global processed_columns
    printed_errors = set()
    # Map source columns to fields
    field_columns = get_source_columns(source_sheet)
    #print(f"field_columns: {field_columns}, at page: {source_sheet.title}")

    # Map target columns to fields
    target_columns = get_target_columns(target_sheet)
    #print(f"target_columns: {target_columns} at page: {target_sheet.title}")

    for row_idx in range(2, source_sheet.max_row + 1):  # Skip header
        target_row = target_sheet.max_row + 1
        for (category, attribute), ((source_sheet_name, source_column)) in processed_columns.items():
            if (category == target_sheet.title) and (source_sheet_name == source_sheet.title):
                try:
                    target_col = target_columns[attribute]
                    source_col = field_columns[source_column]
                    source_cell = source_sheet.cell(row=row_idx, column=source_col)
                    target_cell = target_sheet.cell(row=target_row, column=target_col)
                    target_cell.value = source_cell.value
                except Exception as e:
                    key = (row_idx, attribute)
                    if key not in printed_errors:
                        print(f"Skipping error at row {row_idx}, attribute '{attribute}': {e}")
                        printed_errors.add(key)
                    continue

async def process_workbooks(source_wb, target_wb):
    try:
        # Get all data sheets from both workbooks
        source_sheets = find_data_sheets(source_wb)
        target_sheets = find_data_sheets(target_wb)

        await process_sheets(source_wb)
        
        # Process each corresponding sheet pair
        for source_sheet in source_sheets:
            for target_sheet in target_sheets:
                copy_data_to_target(source_sheet, target_sheet)
                        
    except Exception as e:
        print(f"Error processing data: {str(e)}")
        raise

def create_report(file_format):
    global mappings_attribute
    global processed_columns
    file_sheet = file_format["Báo cáo"]
    row_idx = file_sheet.max_row + 1
    for (category, attribute), ((source_sheet_name, source_column)) in processed_columns.items():
        # Instead of incrementing row_idx here, do it after actual processing
        file_sheet.cell(row=row_idx, column=2, value=category)          
        file_sheet.cell(row=row_idx, column=1, value=attribute)
        
        # Now increment row_idx after processing the row
        row_idx += 1


async def mapping(filepath, formated):
    try:
        #for filepath in filepaths:
        file_khach_hang = load_workbook(filepath, data_only=False)
        file_format = load_workbook(formated)

        await process_workbooks(file_khach_hang, file_format)
            
        # Save and return the formatted file path

        file_format.save(formated)

        create_report(file_format)
        
        file_format.save(formated)

        return formated

    except Exception as e:
        print(f"Error: {str(e)}")
        raise

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Map data from source Excel to formatted Excel based on JSON attributes")
    parser.add_argument("temp_file_path", help="Path to the source Excel file")
    parser.add_argument("temp_file_form_path", help="Path to the formatted Excel file")
    args = parser.parse_args()    

    try:
        asyncio.run(mapping(args.temp_file_path, args.temp_file_form_path))
    except Exception as e:
        print(f"Error during processing: {str(e)}")
        sys.exit(1)
