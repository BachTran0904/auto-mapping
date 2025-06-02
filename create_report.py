from openpyxl import load_workbook

def create_report(formated, report_filename):
    # Load workbook and sheet
    formated_file = load_workbook(formated)
    report_sheet = formated_file["Báo cáo"]

    load_report_filename = load_workbook(report_filename)
    detail_sheet = load_report_filename["Chi tiết"]

    # Loop through rows in report sheet, skipping header row 1
    for report_row_idx in range(2, report_sheet.max_row + 1):
        attribute_value = report_sheet.cell(row=report_row_idx, column=1).value
        category_value = report_sheet.cell(row=report_row_idx, column=2).value

        # Loop through detail sheet rows to find a match
        for detail_row_idx in range(2, detail_sheet.max_row + 1):
            target_attribute_value = detail_sheet.cell(row=detail_row_idx, column=1).value
            target_category_value = detail_sheet.cell(row=detail_row_idx, column=2).value

            if (attribute_value == target_attribute_value) and (category_value == target_category_value):
                detail_sheet.cell(row=detail_row_idx, column=3).value = True
                break

    load_report_filename.save(report_filename)        
    return report_filename
